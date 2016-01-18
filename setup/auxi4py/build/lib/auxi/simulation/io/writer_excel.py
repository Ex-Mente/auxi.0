from auxi.simulation.io.writer import *
from openpyxl import Workbook
from auxi.simulation.path_engine import *
from auxi.simulation.investigation import *


class Writer_Excel(Writer):
    def __init__(self, file_path):
        Writer.__init__(self, file_path)

    def create_file_from_investigation(self, investigation):
        workbook = Workbook()
        sheet_ix = 0
        column_headers = ['Segment', 'Parameter', 'Description', 'Units',
                          'Value']

        for sc_n, sc_val in investigation.scenarios.items():
            row_ix = 1
            ws = workbook.create_sheet(title='Parameters(%s)' % sc_n,
                                       index=sheet_ix)
            for ix, col in enumerate(column_headers):
                ws.cell(row=row_ix, column=ix+1).value = col
            sheet_ix += 1
            for seg_n, seg_val in sc_val.segments.items():
                for p_n, p_val in seg_val.parameters.items():
                    row_ix += 1
                    default_val = str(p_val.default_value)
                    ws.cell(row=row_ix, column=1).value = seg_val.path
                    ws.cell(row=row_ix, column=2).value = p_val.path
                    ws.cell(row=row_ix, column=3).value = p_val.description
                    ws.cell(row=row_ix, column=4).value = p_val.units
                    ws.cell(row=row_ix, column=5).value = default_val

        #df.to_excel(file_path, sc_n, header=False, index=True)
        workbook.save(self.file_path)
