from auxi.simulation.io.reader import *
from openpyxl import load_workbook
import pandas as pd
from auxi.simulation.path_engine import *
from auxi.simulation.investigation import *
from dateutil.parser import *


class Reader_Excel(Reader):
    def __init__(self, file_path):
        Reader.__init__(self, file_path)

    def update_scenario_from_file(self, investigation, scenario_name):
        workbook = load_workbook(self.file_path)

        scen_name = 'Parameters('+scenario_name+')'
        worksheet = workbook.get_sheet_by_name(scen_name)
        row_count = worksheet.max_row
        column_headers = ['Segment', 'Parameter',
                          'Description', 'Units', 'Value']
        dataframe = pd.DataFrame(columns=column_headers)

        # Clear all the current parameters in the scenario's segments.
        for seg_name, seg in investigation.scenarios[scenario_name].segments.items():
            seg.parameters.clear()

        # Update the scenario's segments with the parameters in the excel file.
        for count in range(2, row_count+1):
            segment_name = worksheet.cell(row=count, column=1).value
            segment_name = segment_name[segment_name.find("::")+2:]
            dataframe.loc[count, 'Segment'] = segment_name
            dataframe.loc[count, 'Parameter'] = worksheet.cell(row=count,
                                                               column=2).value
            dataframe.loc[count, 'Description'] = worksheet.cell(row=count,
                                                                 column=3).value
            dataframe.loc[count, 'Units'] = worksheet.cell(row=count,
                                                           column=4).value
            dataframe.loc[count, 'Value'] = worksheet.cell(row=count,
                                                           column=5).value

            excel_file_path = './' + self.file_path
            time_set_df = pd.read_excel(excel_file_path, scen_name)
            col_ix = 5
            number_of_settings = int((len(time_set_df.columns) - col_ix)/2)
            time_column_name = "DateTime_TimeStep_Interval"
            value_column_name = "Value"
            for j in range(1, number_of_settings+1):
                time_col = time_column_name+"."+str(j)
                val_col = value_column_name+"."+str(j)
                dataframe.loc[count, time_col] = worksheet.cell(
                    row=count, column=col_ix+(j*2)-1).value
                dataframe.loc[count, val_col] = worksheet.cell(
                    row=count, column=col_ix+(j*2)).value

            # Create the scenario in investigation
            if scenario_name in investigation.scenarios:
                scenario = investigation.scenarios[scenario_name]
            else:
                raise ValueError("The scenario '%s' does not exist in the investigation" % scenario_name)

            for param in dataframe.index:
                segment_name = dataframe.at[param, 'Segment']
                param_name = dataframe.at[param, 'Parameter']
                param_units = dataframe.at[param, 'Units']
                param_default_value = dataframe.at[param, 'Value']
                param_desc = dataframe.at[param, 'Description']
                # Don't create a new segment if the segment exists.
                if segment_name in scenario.segments:
                    segment = scenario[segment_name]
                else:
                    raise ValueError("The segment '%s' does not exist in the scenario '%s'." % (segment_name, scenario_name))

                #Create parameter in the investigation
                param_source = get_instance_value(segment.auxi_object,
                                                  param_name)
                new_param = segment.create_parameter(param_name,
                                                     param_source,
                                                     param_units,
                                                     param_default_value,
                                                     param_desc)

                #set_instance_value(segment.auxi_object,
                #                   param_name, param_default_value)

                for j in range(1, number_of_settings+1):
                    time_col = time_column_name+"."+str(j)
                    val_col = value_column_name+"."+str(j)
                    time_val = dataframe.at[param, time_col]
                    val = dataframe.at[param, val_col]
                    if time_val == time_val and val == val:
                        try:
                            dt = parse(time_val)
                            new_param.value_dict_dates[dt] = val
                        except AttributeError:
                            ix = eval(str(time_val))
                            if isinstance(ix, float):
                                new_param.value_dict_intervals[ix] = int(val)
